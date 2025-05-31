import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook
from modules import database
from datetime import datetime
from openpyxl.styles import Font, Alignment

class ReportFrame(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.data = []
        self.build_ui()
        self.load_data()

    def build_ui(self):
        columns = ["Item Name", "Item Size", "Item GSM", "Item BF", "Item Reels", "Closing Stock"]
        self.tree = ttk.Treeview(self, columns=columns, show="headings", height=15)
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)
        self.tree.pack(fill="x", padx=10, pady=10)

        tk.Button(self, text="Export to Excel", command=self.export_to_excel).pack(pady=10)

    def load_data(self):
        self.data = database.get_stock_report()
        for row in self.data:
            self.tree.insert("", "end", values=row)

    def export_to_excel(self):
        if not self.data:
            messagebox.showinfo("No Data", "No stock data to export.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialdir=os.path.abspath("exports"),
            title="Save Excel File"
        )
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "TodayStock"
        today = datetime.now().strftime("%Y-%m-%d")
        ws.append([f"Closing Stock Report - {today}"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Item Name", "Item Size", "Item GSM", "Item BF", "No. of Reels", "Closing Stock"]
        ws.append(headers)

        column_widths = [25, 15, 15, 15, 18, 10]
        for i, width in enumerate(column_widths):
            col_letter = chr(65 + i)  # Convert index to Excel column letter (A, B, ...)
            ws.column_dimensions[col_letter].width = width

        for row in self.data:
            ws.append(row)

        wb.save(file_path)
        messagebox.showinfo("Exported", f"Stock exported to:\n{file_path}")
